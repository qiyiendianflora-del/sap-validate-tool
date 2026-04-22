#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SAP凭证校验工具 v3.0 - PyQt5 GUI（橙色主题 + 卡片式布局 + 字体缩放）
打包：pyinstaller --onefile --windowed --name SAP凭证校验工具 validate_gui.py
"""
import sys, os, json, shutil
from datetime import datetime
from decimal import Decimal, InvalidOperation
from collections import defaultdict

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QProgressBar, QMessageBox, QGroupBox, QAbstractItemView,
    QTabBar, QFrame, QGraphicsDropShadowEffect, QStackedWidget, QToolTip,
    QShortcut, QMenuBar, QMenu, QAction, QDialog, QSlider, QDialogButtonBox,
    QSizePolicy, QGridLayout
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QPoint
from PyQt5.QtGui import QFont, QColor, QDragEnterEvent, QDropEvent, QKeySequence

import openpyxl
from openpyxl.styles import PatternFill, Font as XlFont, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# ━━━━━━━ 常量 ━━━━━━━
APP_NAME="SAP凭证校验工具"; VERSION="v3.9.3"; DATA_ROW=4
CONFIG_FILE=os.path.join(os.path.expanduser("~"),".sap_validate_config.json")
CACHE_FILE=os.path.join(os.path.expanduser("~"),".sap_validate_cache.json")
CACHE_DIR=os.path.join(os.path.expanduser("~"),".sap_validate_files")
os.makedirs(CACHE_DIR, exist_ok=True)

DEBIT_CODES={'01','21','40'}; CREDIT_CODES={'11','31','50'}; ALL_CODES=DEBIT_CODES|CREDIT_CODES
RECON_CODES={'D':{'01','11'},'K':{'21','31'},None:{'40','50'},'':{'40','50'}}
RECON_LABELS={'D':'客户统驭','K':'供应商统驭',None:'普通科目','':'普通科目'}
HARDCODED_RULES=[(1002010001,1002019999,[('AD','现金流量码'),('X','利润中心')])]
FORBIDDEN_COLS_NON6=[
    ('AF','客户'),('AG','供应商'),('AH','生产（物料）'),('AI','流量类型'),
    ('AJ','渠道'),('AK','一级费用'),('AL','二级费用'),('AM','城市'),
    ('AN','项目'),('AO','销售部门'),('AP','产品'),('AQ','提报人'),
    ('AR','提报部门'),('AS','报销人'),('AT','销售大区'),('AU','销售员')]

A_IDX=0; B_IDX=1; J_IDX=column_index_from_string('J')-1; M_IDX=column_index_from_string('M')-1
O_IDX=column_index_from_string('O')-1; AK_IDX=column_index_from_string('AK')-1; AL_IDX=column_index_from_string('AL')-1

# ── 配色方案（橙色主题） ──
C_PRIMARY   = '#FF8C00'   # 主色/按钮
C_PRIMARY_H = '#F07800'   # hover
C_PRIMARY_L = '#FFF4E6'   # 浅底色
C_BG        = '#FAFAF8'   # 背景
C_CARD      = '#FFFFFF'   # 卡片
C_TEXT      = '#1C1917'   # 文字主色
C_TEXT_SEC  = '#78716C'   # 文字次色
C_TEXT_AUX  = '#A8A29E'   # 文字辅助
C_SUCCESS   = '#16A34A'   # 成功
C_ERROR     = '#DC2626'   # 错误
C_WARN      = '#FF8C00'   # 警告
C_BORDER    = '#E7E5E4'   # 边框
C_BORDER_L  = '#F5F5F4'   # 浅边框
C_DROP_BG   = '#EBF5FF'   # 拖拽区背景
C_DROP_BD   = '#B0D4F1'   # 拖拽区虚线
C_ACCENT    = '#E65100'   # 强调色

HEADER_ROW1=['凭证序号','公司代码','凭证类型','凭证日期','过账日期','货币','汇率','凭证抬头文本','参照','记帐代码','屏幕科目','特别总帐标识','总账科目','事物类型','凭证货币金额','本币金额','成本中心编号','订单号','基准日期','参考','分配','项目文本','反记账标识','利润中心','数量','单位','付款条件','冻结付款','功能范围','现金流量码','贸易伙伴','客户','供应商','生产（物料）','流量类型','渠道','一级费用','二级费用','城市','项目','销售部门','产品','提报人','提报部门','报销人','销售大区','销售员','银行流水号','物料','OA单据号','销售合同号']
HEADER_ROW2=['LINEID','BUKRS','BLART','BLDAT','BUDAT','WAERS','KURSF','BKTXT','XBLNR','BSCHL','NEWKO','UMSKZ','HKONT','BEWAR','WRBTR','DMBTR','KOSTL','AUFNR','ZFBDT','XREF1','ZUONR','SGTXT','XNEGP','PRCTR','MENGE','MEINS','ZTERM','ZLSPR','FKBER','ZZCASHFLOW','RASSC','KUNNR','WWVEN','artnr','WWLL','WWOFR','WWFY1','WWFY2','WWCS','WWXM','WWXSB','WWPRC','WWTBR','WWTBM','WWBXR','WWXSQ','WWXSY','XREF3','MATNR','XREF1_HD','XREF2_HD']
HEADER_ROW3=[10,4,2,'8(YYYYMMDD)','8(YYYYMMDD)',5,'(9,5)',25,16,2,10,1,10,3,'(13,2)','(13,2)',10,12,'8(YYYYMMDD)',12,18,50,1,10,'(13,3)',3,4,1,4,5,5,6,6,13,18,18,18,18,18,18,18,18,18,10,18,18,18,20,18,20,20]
NUM_COLS=len(HEADER_ROW1)
COL_LETTERS=[get_column_letter(i+1) for i in range(NUM_COLS)]
RESULT_COL=0
DATA_COL_OFFSET=1
GRID_HEADER_ROWS=1  # 只保留中文列名一行

# ━━━━━━━ 字体体系（基准值 + 缩放） ━━━━━━━
FONT_FAMILY = "微软雅黑"
BASE_TITLE  = 18
BASE_H2     = 13
BASE_BODY   = 12
BASE_SMALL  = 8
BASE_GRID   = 9
BASE_GRID_H = 17
BASE_BTN    = 12
BASE_STAT_NUM = 22

# 全局缩放比例（运行时动态修改）
_font_scale = 1.0

def fs(base):
    """根据缩放比例计算实际字号"""
    global _font_scale
    return max(8, int(base * _font_scale + 0.5))

def FONT_TITLE():  return fs(BASE_TITLE)
def FONT_H2():     return fs(BASE_H2)
def FONT_BODY():   return fs(BASE_BODY)
def FONT_SMALL():  return fs(BASE_SMALL)
def FONT_GRID():   return fs(BASE_GRID)
def FONT_GRID_H(): return fs(BASE_GRID_H)
def FONT_BTN():    return fs(BASE_BTN)
def FONT_STAT():   return fs(BASE_STAT_NUM)

# ━━━━━━━ 配置持久化 ━━━━━━━
def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE,'r',encoding='utf-8') as f: return json.load(f)
        except: pass
    return {}
def save_config(cfg):
    try:
        with open(CONFIG_FILE,'w',encoding='utf-8') as f: json.dump(cfg,f,ensure_ascii=False,indent=2)
    except: pass
def load_mapping_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE,'r',encoding='utf-8') as f: return json.load(f)
        except: pass
    return None
def save_mapping_cache(data):
    try:
        with open(CACHE_FILE,'w',encoding='utf-8') as f: json.dump(data,f,ensure_ascii=False,indent=2)
    except: pass

# ━━━━━━━ 校验逻辑（与 v2.1 相同） ━━━━━━━
def col_letter_to_idx(letter): return column_index_from_string(letter)-1
def get_hardcoded_required(acct_str):
    try: acct_int=int(acct_str)
    except: return None
    for s,e,cols in HARDCODED_RULES:
        if s<=acct_int<=e: return cols
    return None
def load_rule_table(rule_file):
    wb=openpyxl.load_workbook(rule_file);ws=wb.active;rows=list(ws.iter_rows(values_only=True))
    fn,ci=rows[0],rows[1]; rm,rcm={},{}
    for row in rows[2:]:
        a=row[0]
        if a is None: continue
        a_s=str(int(a)) if isinstance(a,float) else str(a).strip()
        rc=row[2] if len(row)>2 else None
        if rc: rc=str(rc).strip()
        rcm[a_s]=rc if rc else None
        req=[]
        for f,c,v in zip(fn,ci,row):
            if c and c!='M' and v=='√': req.append((c,f))
        rm[a_s]=req
    return rm,rcm
def load_mapping_table(mf):
    wb=openpyxl.load_workbook(mf);ws=wb.active;rows=list(ws.iter_rows(values_only=True))
    mp=defaultdict(set)
    for row in rows[1:]:
        if row[0] is None: continue
        a=str(int(row[0])) if isinstance(row[0],float) else str(row[0]).strip()
        f1=str(int(row[1])) if isinstance(row[1],float) else str(row[1]).strip() if row[1] else ''
        f2=str(int(row[2])) if isinstance(row[2],float) else str(row[2]).strip() if row[2] else ''
        mp[a].add((f1,f2))
    return {k:list(v) for k,v in mp.items()}
def safe_decimal(val):
    if val is None: return None
    try:
        # 去除千分位逗号（如 "4,071.52" -> "4071.52"）
        cleaned = str(val).strip().replace(',', '')
        return Decimal(cleaned)
    except: return None
def check_decimal_places(val):
    d=safe_decimal(val)
    if d is None: return True
    s,dg,exp=d.as_tuple()
    return abs(exp)<=2 if exp<0 else True
def _norm(val):
    if val is None: return ''
    if isinstance(val,float): return str(int(val))
    return str(val).strip()

def run_validate_data(all_rows, rule_map, recon_map, mapping_data=None, progress_cb=None):
    mapping_sets={}
    if mapping_data:
        for a,combos in mapping_data.items(): mapping_sets[a]=set(tuple(c) for c in combos)
    total=len(all_rows)
    if total==0: return {},[]
    errors=defaultdict(list); warnings=[]
    for idx,(excel_row,rv) in enumerate(all_rows.items()):
        if progress_cb and idx%50==0: progress_cb(int(idx/total*55))
        acct_val=rv[M_IDX] if M_IDX<len(rv) else None
        if acct_val is None: continue
        acct_str=_norm(acct_val)
        hc=get_hardcoded_required(acct_str); is_hc=hc is not None
        if is_hc: req_cols=hc
        elif acct_str in rule_map: req_cols=rule_map[acct_str]
        else: warnings.append((excel_row,acct_str,'科目不在规则表')); req_cols=[]
        for cl,fn in req_cols:
            ci=col_letter_to_idx(cl); v=rv[ci] if ci<len(rv) else None
            if v is None or str(v).strip()=='':
                errors[excel_row].append(('必输项为空',[cl],f'缺失必输项：{cl}({fn})'))
        o_val=rv[O_IDX] if O_IDX<len(rv) else None
        if o_val is None or str(o_val).strip()=='':
            errors[excel_row].append(('金额为空',['O'],'凭证货币金额(O列)为空'))
        elif not check_decimal_places(o_val):
            errors[excel_row].append(('金额格式错误',['O'],f'金额小数位超过2位：{o_val}'))
        j_val=rv[J_IDX] if J_IDX<len(rv) else None; j_str=_norm(j_val)
        if len(j_str)==1: j_str='0'+j_str
        if not is_hc and acct_str in recon_map:
            rt=recon_map[acct_str]; rk=''
            if rt:
                ru=str(rt).strip().upper()
                if '客户' in str(rt) or ru=='D': rk='D'
                elif '供应' in str(rt) or ru=='K': rk='K'
            allowed=RECON_CODES.get(rk,set()); label=RECON_LABELS.get(rk,'普通科目')
            if j_str and j_str not in allowed:
                errors[excel_row].append(('记账码不匹配',['J'],f'统驭科目为{label}({rk or "空"})，应为{"/".join(sorted(allowed))}，实际{j_str}'))
        if j_str and j_str not in ALL_CODES:
            warnings.append((excel_row,j_str,'记账码不在标准范围'))
        if not acct_str.startswith('6'):
            for cl,fn in FORBIDDEN_COLS_NON6:
                ci=col_letter_to_idx(cl); v=rv[ci] if ci<len(rv) else None
                if v is not None and str(v).strip()!='':
                    errors[excel_row].append(('禁填字段有值',[cl],f'该科目非损益科目（非6开头），无需填写{fn}信息，请清空"{fn}（{cl}列）"字段'))
        if mapping_sets and acct_str in mapping_sets:
            ak_val=_norm(rv[AK_IDX] if AK_IDX<len(rv) else None)
            al_val=_norm(rv[AL_IDX] if AL_IDX<len(rv) else None)
            if ak_val and al_val:
                if (ak_val,al_val) not in mapping_sets[acct_str]:
                    errors[excel_row].append(('费用类别不匹配',['AK','AL'],f'科目{acct_str}下，一级费用{ak_val}+二级费用{al_val}组合不在配置表中'))
    if progress_cb: progress_cb(65)
    groups=defaultdict(list)
    for r,rv in all_rows.items():
        a=rv[A_IDX] if A_IDX<len(rv) else None
        if a is None: continue
        groups[(_norm(a),_norm(rv[B_IDX] if B_IDX<len(rv) else None))].append(r)
    if progress_cb: progress_cb(75)
    for (lid,bk),rlist in groups.items():
        ds,cs=Decimal('0'),Decimal('0')
        for r in rlist:
            rv=all_rows[r]; ov=rv[O_IDX] if O_IDX<len(rv) else None
            jv=rv[J_IDX] if J_IDX<len(rv) else None; js=_norm(jv)
            if len(js)==1: js='0'+js
            amt=safe_decimal(ov)
            if amt is None: continue
            if js in DEBIT_CODES: ds+=amt
            elif js in CREDIT_CODES: cs+=amt
        diff=ds-cs
        if diff!=0:
            msg=f'借贷不平衡：凭证{lid}/公司{bk}，借方{ds} 贷方{cs} 差额{diff}'
            for r in rlist: errors[r].append(('借贷不平衡',[],msg))
    if progress_cb: progress_cb(90)
    return errors,warnings

def run_validate(b_file, rule_file, mapping_data=None, progress_cb=None):
    rule_map,recon_map=load_rule_table(rule_file)
    wb_b=openpyxl.load_workbook(b_file);ws_b=wb_b.active;mc=ws_b.max_column
    total=ws_b.max_row-DATA_ROW+1
    if total<=0: return {},[], openpyxl.Workbook(),ws_b,rule_map,0
    all_rows={}
    for r in range(DATA_ROW,ws_b.max_row+1):
        all_rows[r]=[ws_b.cell(row=r,column=c+1).value for c in range(mc)]
    errors,warnings=run_validate_data(all_rows,rule_map,recon_map,mapping_data,progress_cb)
    out_wb=build_report(errors,warnings,wb_b,ws_b,rule_map,total)
    if progress_cb: progress_cb(100)
    return errors,warnings,out_wb,ws_b,rule_map,total

def build_report(errors,warnings,wb_b,ws_b,rule_map,total):
    RED=PatternFill(start_color="FF4444",end_color="FF4444",fill_type="solid")
    YELLOW=PatternFill(start_color="FFF2CC",end_color="FFF2CC",fill_type="solid")
    ORANGE=PatternFill(start_color="FFE0B2",end_color="FFE0B2",fill_type="solid")
    HDR=PatternFill(start_color="E65100",end_color="E65100",fill_type="solid")
    GRAY=PatternFill(start_color="F2F2F2",end_color="F2F2F2",fill_type="solid")
    WHITE=PatternFill(start_color="FFFFFF",end_color="FFFFFF",fill_type="solid")
    thin=Side(style="thin",color="DDDDDD");bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    out_wb=openpyxl.Workbook();ws_out=out_wb.active;ws_out.title="校验结果"
    ecells=set()
    for r,el in errors.items():
        for _,cols,_ in el:
            for cl in cols: ecells.add((r,cl))
    wa={w[0]:w[1] for w in warnings if w[2]=='科目不在规则表'}
    mc=ws_b.max_column;ec=mc+1
    for r in range(1,ws_b.max_row+1):
        for c in range(1,mc+1):
            v=ws_b.cell(row=r,column=c).value;d=ws_out.cell(row=r,column=c,value=v)
            d.border=bdr;d.alignment=Alignment(vertical="center");cl=get_column_letter(c)
            if r<=3:
                d.fill=HDR if r==1 else GRAY
                d.font=XlFont(name="微软雅黑",bold=(r==1),color="FFFFFF" if r==1 else "333333",size=10)
            elif r in errors:
                d.fill=RED if (r,cl) in ecells else YELLOW
                d.font=XlFont(name="微软雅黑",color="FFFFFF" if (r,cl) in ecells else "000000",size=10)
            elif r in wa: d.fill=ORANGE;d.font=XlFont(name="微软雅黑",size=10)
            else: d.fill=WHITE;d.font=XlFont(name="微软雅黑",size=10)
        if r==1:
            c2=ws_out.cell(row=1,column=ec,value="错误说明");c2.fill=HDR
            c2.font=XlFont(name="微软雅黑",bold=True,color="FFFFFF",size=10);c2.border=bdr
        if r in errors:
            c2=ws_out.cell(row=r,column=ec,value=" | ".join([e[2] for e in errors[r]]))
            c2.fill=YELLOW;c2.font=XlFont(name="微软雅黑",color="CC0000",size=10);c2.border=bdr
        elif r in wa:
            c2=ws_out.cell(row=r,column=ec,value=f"⚠️ 科目{wa[r]}不在规则表中")
            c2.fill=ORANGE;c2.font=XlFont(name="微软雅黑",color="885500",size=10);c2.border=bdr
    for c in range(1,mc+1): ws_out.column_dimensions[get_column_letter(c)].width=14
    ws_out.column_dimensions[get_column_letter(ec)].width=65;ws_out.freeze_panes=f"A{DATA_ROW}"
    ws2=out_wb.create_sheet("错误汇总")
    for i,h in enumerate(["行号","总账科目","错误类型","错误详情"],1):
        c2=ws2.cell(row=1,column=i,value=h);c2.fill=HDR;c2.font=XlFont(name="微软雅黑",bold=True,color="FFFFFF",size=11)
        c2.alignment=Alignment(horizontal="center",vertical="center");c2.border=bdr
    ws2.column_dimensions['A'].width=8;ws2.column_dimensions['B'].width=16
    ws2.column_dimensions['C'].width=18;ws2.column_dimensions['D'].width=65
    ri=2
    for r in sorted(errors.keys()):
        rv=[ws_b.cell(row=r,column=c+1).value for c in range(ws_b.max_column)]
        a_s=_norm(rv[M_IDX] if M_IDX<len(rv) else '')
        for et,cols,msg in errors[r]:
            for i,v in enumerate([r,a_s,f"❌ {et}",msg],1):
                c2=ws2.cell(row=ri,column=i,value=v)
                c2.fill=PatternFill(start_color="FFF2CC",end_color="FFF2CC",fill_type="solid")
                c2.font=XlFont(name="微软雅黑",size=10);c2.alignment=Alignment(vertical="center",wrap_text=(i==4));c2.border=bdr
            ri+=1
    return out_wb

# ━━━━━━━ 后台线程 ━━━━━━━
class ValidateWorker(QThread):
    progress=pyqtSignal(int);finished=pyqtSignal(object);error=pyqtSignal(str)
    def __init__(self,b,r,m=None): super().__init__();self.b=b;self.r=r;self.m=m
    def run(self):
        try:
            e,w,wb,ws,rm,t=run_validate(self.b,self.r,self.m,lambda v:self.progress.emit(v))
            self.finished.emit({"errors":e,"warnings":w,"out_wb":wb,"ws":ws,"total":t})
        except: import traceback;self.error.emit(traceback.format_exc())

class GridValidateWorker(QThread):
    progress=pyqtSignal(int);finished=pyqtSignal(object);error=pyqtSignal(str)
    def __init__(self,all_rows,rule_file,mapping_data=None):
        super().__init__();self.all_rows=all_rows;self.rule_file=rule_file;self.mapping_data=mapping_data
    def run(self):
        try:
            rm,rcm=load_rule_table(self.rule_file)
            e,w=run_validate_data(self.all_rows,rm,rcm,self.mapping_data,lambda v:self.progress.emit(v))
            self.finished.emit({"errors":e,"warnings":w,"total":len(self.all_rows)})
        except: import traceback;self.error.emit(traceback.format_exc())

# ━━━━━━━ 支持多行粘贴的 QTableWidget ━━━━━━━
class PasteableTable(QTableWidget):
    """重写 keyPressEvent 支持 Ctrl+V 多行多列粘贴"""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def keyPressEvent(self, event):
        if event.matches(QKeySequence.Paste):
            self._paste_from_clipboard()
        else:
            super().keyPressEvent(event)

    def _paste_from_clipboard(self):
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        if not text:
            return
        lines = text.rstrip('\n').split('\n')
        rows = [line.split('\t') for line in lines]

        sel = self.selectedItems()
        if sel:
            start_row = sel[0].row()
            start_col = sel[0].column()
        else:
            start_row = self.currentRow()
            start_col = self.currentColumn()

        if start_row < GRID_HEADER_ROWS:
            start_row = GRID_HEADER_ROWS
        if start_col < DATA_COL_OFFSET:
            start_col = DATA_COL_OFFSET

        needed = start_row + len(rows)
        if needed > self.rowCount():
            self.setRowCount(needed + 50)

        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                r = start_row + i
                c = start_col + j
                if c >= self.columnCount():
                    break
                if r < GRID_HEADER_ROWS:
                    continue
                item = self.item(r, c)
                if item is None:
                    item = QTableWidgetItem()
                    self.setItem(r, c, item)
                if item.flags() & Qt.ItemIsEditable:
                    item.setText(val.strip())
                elif not (item.flags() & Qt.ItemIsEnabled):
                    pass
                else:
                    item.setText(val.strip())

        self.viewport().update()


# ━━━━━━━ 样式工具函数 ━━━━━━━
def _orange_btn_style(sz=None):
    """橙色渐变按钮样式"""
    if sz is None:
        sz = FONT_BTN()
    return f"""
        QPushButton {{
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #FFA033, stop:1 #FF8C00);
            color: #FFFFFF; border-radius: 10px;
            padding: 8px 20px; font-size: {sz}px; font-weight: bold;
            border: none;
        }}
        QPushButton:hover {{
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #FF9520, stop:1 #F07800);
        }}
        QPushButton:pressed {{
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #E87000, stop:1 #D06000);
        }}
        QPushButton:disabled {{ background: #ccc; color: #888; }}
    """

def _secondary_btn_style(sz=None):
    """白底灰边框按钮样式"""
    if sz is None:
        sz = FONT_BTN()
    return f"""
        QPushButton {{
            background: #FFFFFF; color: {C_TEXT_SEC};
            border: 1px solid {C_BORDER}; border-radius: 10px;
            padding: 8px 20px; font-size: {sz}px; font-weight: 500;
        }}
        QPushButton:hover {{
            background: {C_BG}; color: {C_TEXT};
        }}
        QPushButton:disabled {{ background: #f5f5f5; color: #bbb; border-color: #e0e0e0; }}
    """

def _danger_btn_style(sz=None):
    """危险操作按钮样式"""
    if sz is None:
        sz = FONT_BTN()
    return f"""
        QPushButton {{
            background: #FFFFFF; color: {C_ERROR};
            border: 1px solid #FDE68A; border-radius: 10px;
            padding: 8px 20px; font-size: {sz}px; font-weight: 500;
        }}
        QPushButton:hover {{
            background: #FEF3C7; border-color: #E8A87C;
        }}
    """


# ━━━━━━━ 自定义控件 ━━━━━━━

class StatCard(QFrame):
    """统计卡片：图标 + 大数字 + 标签"""
    def __init__(self, icon, label, color, num_color):
        super().__init__()
        self._icon = icon
        self._label_text = label
        self._color = color
        self._num_color = num_color
        self._value = "—"

        self.setStyleSheet(f"""
            StatCard {{
                background: {C_CARD}; border: 1px solid {C_BORDER};
                border-radius: 10px;
            }}
        """)
        ly = QHBoxLayout(self)
        ly.setContentsMargins(16, 12, 16, 12)
        ly.setSpacing(12)

        # 图标容器
        self.icon_frame = QFrame()
        self.icon_frame.setFixedSize(44, 44)
        self.icon_frame.setStyleSheet(f"""
            QFrame {{
                background: {color}; border-radius: 10px;
            }}
        """)
        icon_ly = QHBoxLayout(self.icon_frame)
        icon_ly.setContentsMargins(0, 0, 0, 0)
        self.icon_lbl = QLabel(icon)
        self.icon_lbl.setAlignment(Qt.AlignCenter)
        self.icon_lbl.setStyleSheet("font-size: 20px; background: transparent; border: none;")
        icon_ly.addWidget(self.icon_lbl)
        ly.addWidget(self.icon_frame)

        # 数字和标签
        info_ly = QVBoxLayout()
        info_ly.setSpacing(2)
        self.num_lbl = QLabel("—")
        self.num_lbl.setStyleSheet(f"""
            font-size: {FONT_STAT()}px; font-weight: bold;
            color: {num_color}; background: transparent; border: none;
        """)
        self.name_lbl = QLabel(label)
        self.name_lbl.setStyleSheet(f"""
            font-size: {FONT_SMALL()}px; font-weight: 500;
            color: {C_TEXT_AUX}; background: transparent; border: none;
        """)
        info_ly.addWidget(self.num_lbl)
        info_ly.addWidget(self.name_lbl)
        ly.addLayout(info_ly)
        ly.addStretch()

    def set_value(self, val):
        self._value = str(val)
        self.num_lbl.setText(self._value)

    def refresh_style(self):
        global _font_scale
        icon_sz = max(36, int(44 * _font_scale + 0.5))
        self.icon_frame.setFixedSize(icon_sz, icon_sz)
        self.num_lbl.setStyleSheet(f"""
            font-size: {FONT_STAT()}px; font-weight: bold;
            color: {self._num_color}; background: transparent; border: none;
        """)
        self.name_lbl.setStyleSheet(f"""
            font-size: {FONT_SMALL()}px; font-weight: 500;
            color: {C_TEXT_AUX}; background: transparent; border: none;
        """)


class StatRow(QFrame):
    """统计卡片行：4 个 StatCard 横排"""
    def __init__(self):
        super().__init__()
        ly = QHBoxLayout(self)
        ly.setContentsMargins(0, 0, 0, 0)
        ly.setSpacing(16)

        self.cards = {}
        configs = [
            ("📊", "总行数", C_PRIMARY_L, C_PRIMARY, "total"),
            ("✅", "通过", "#F0FDF4", C_SUCCESS, "ok"),
            ("❌", "错误", "#FEF3C7", C_ERROR, "err"),
            ("⚠️", "警告", "#FFFBEB", C_WARN, "warn"),
        ]
        for icon, label, bg, num_color, key in configs:
            card = StatCard(icon, label, bg, num_color)
            ly.addWidget(card, 1)
            self.cards[key] = card

    def set_values(self, total="—", ok="—", err="—", warn="—"):
        self.cards["total"].set_value(total)
        self.cards["ok"].set_value(ok)
        self.cards["err"].set_value(err)
        self.cards["warn"].set_value(warn)

    def refresh_style(self):
        for card in self.cards.values():
            card.refresh_style()


class DropArea(QLabel):
    """拖拽区域：浅蓝色背景 + 虚线边框"""
    file_dropped = pyqtSignal(str)

    def __init__(self, text="将文件拖拽到此处，或点击上方按钮选择", min_h=160):
        super().__init__(text)
        self.setAcceptDrops(True)
        self.setAlignment(Qt.AlignCenter)
        self._is_loaded = False
        self._loaded_fn = ""
        self._min_h = min_h
        self.setMinimumHeight(min_h)
        self._apply_idle()

    def _apply_idle(self):
        self._is_loaded = False
        self.setStyleSheet(f"""
            QLabel {{
                border: 2px dashed {C_DROP_BD};
                border-radius: 10px;
                color: {C_TEXT_AUX};
                background: {C_DROP_BG};
                font-size: {FONT_BODY()}px;
                padding: 8px;
            }}
        """)

    def _apply_loaded(self):
        self._is_loaded = True
        self.setStyleSheet(f"""
            QLabel {{
                border: 2px solid {C_DROP_BD};
                border-radius: 10px;
                color: {C_ACCENT};
                background: {C_DROP_BG};
                font-size: {FONT_BODY()}px;
                font-weight: bold;
                padding: 8px;
            }}
        """)

    def set_loaded(self, fn):
        self._loaded_fn = fn
        self.setText(f"📄 {fn}")
        self._apply_loaded()

    def reset_idle(self, text=None):
        if text:
            self.setText(text)
        else:
            self.setText("将文件拖拽到此处，或点击上方按钮选择")
        self._apply_idle()

    def refresh_style(self):
        global _font_scale
        h = max(self._min_h, int(self._min_h * _font_scale + 0.5))
        self.setMinimumHeight(h)
        if self._is_loaded:
            self._apply_loaded()
        else:
            self._apply_idle()

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()

    def dragLeaveEvent(self, e):
        if not self._is_loaded:
            self._apply_idle()

    def dropEvent(self, e):
        urls = e.mimeData().urls()
        if urls:
            p = urls[0].toLocalFile()
            if p.lower().endswith(('.xlsx', '.xls')):
                self.set_loaded(os.path.basename(p))
                self.file_dropped.emit(p)
            else:
                QMessageBox.warning(self, "格式错误", "请拖入Excel文件")
                self._apply_idle()


class FileBlock(QFrame):
    """文件上传卡片块：标题 + 按钮行 + 拖拽区 + 状态"""
    file_chosen = pyqtSignal(str)

    def __init__(self, title, subtitle="", drop_text="将文件拖拽到此处，或点击上方按钮选择",
                 drop_h=160, show_reload=True):
        super().__init__()
        self._title_text = title
        self.setStyleSheet(f"""
            FileBlock {{
                background: {C_CARD};
                border: 1px solid {C_BORDER};
                border-radius: 10px;
            }}
        """)
        ly = QVBoxLayout(self)
        ly.setContentsMargins(20, 16, 20, 16)
        ly.setSpacing(8)

        # 标题行
        title_row = QHBoxLayout()
        self.title_lbl = QLabel(title)
        self.title_lbl.setStyleSheet(f"""
            font-size: {FONT_H2()}px; font-weight: bold;
            color: {C_PRIMARY}; background: transparent; border: none;
        """)
        title_row.addWidget(self.title_lbl)
        if subtitle:
            self.subtitle_lbl = QLabel(f"· {subtitle}")
            self.subtitle_lbl.setStyleSheet(f"""
                font-size: {FONT_SMALL()}px; font-weight: 400;
                color: {C_TEXT_AUX}; background: transparent; border: none;
            """)
            title_row.addWidget(self.subtitle_lbl)
        else:
            self.subtitle_lbl = None
        title_row.addStretch()
        ly.addLayout(title_row)

        # 按钮行
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        self.choose_btn = QPushButton("📂 选择文件")
        self.choose_btn.setFixedHeight(36)
        self.choose_btn.setStyleSheet(_orange_btn_style())
        self.choose_btn.clicked.connect(self._on_choose)
        btn_row.addWidget(self.choose_btn)

        self.reload_btn = None
        if show_reload:
            self.reload_btn = QPushButton("🔄 重新加载")
            self.reload_btn.setFixedHeight(36)
            self.reload_btn.setStyleSheet(_orange_btn_style())
            btn_row.addWidget(self.reload_btn)

        btn_row.addStretch()
        ly.addLayout(btn_row)

        # 拖拽区
        self.drop = DropArea(drop_text, min_h=drop_h)
        self.drop.file_dropped.connect(lambda p: self.file_chosen.emit(p))
        ly.addWidget(self.drop)

        # 状态标签
        self.status_lbl = QLabel("")
        self.status_lbl.setStyleSheet(f"""
            font-size: {FONT_SMALL()}px; font-weight: 500;
            color: {C_TEXT_AUX}; background: transparent; border: none;
        """)
        ly.addWidget(self.status_lbl)

    def set_status(self, text, ok=True):
        color = C_SUCCESS if ok else C_TEXT_AUX
        self.status_lbl.setText(text)
        self.status_lbl.setStyleSheet(f"""
            font-size: {FONT_SMALL()}px; font-weight: 500;
            color: {color}; background: transparent; border: none;
        """)

    def _on_choose(self):
        p, _ = QFileDialog.getOpenFileName(None, "选择文件", "", "Excel (*.xlsx *.xls)")
        if p:
            self.drop.set_loaded(os.path.basename(p))
            self.file_chosen.emit(p)

    def refresh_style(self):
        self.title_lbl.setStyleSheet(f"""
            font-size: {FONT_H2()}px; font-weight: bold;
            color: {C_PRIMARY}; background: transparent; border: none;
        """)
        if self.subtitle_lbl:
            self.subtitle_lbl.setStyleSheet(f"""
                font-size: {FONT_SMALL()}px; font-weight: 400;
                color: {C_TEXT_AUX}; background: transparent; border: none;
            """)
        self.choose_btn.setStyleSheet(_orange_btn_style())
        if self.reload_btn:
            self.reload_btn.setStyleSheet(_orange_btn_style())
        self.drop.refresh_style()
        # re-apply status style
        cur_text = self.status_lbl.text()
        if cur_text:
            is_ok = "✔" in cur_text or "✅" in cur_text
            self.set_status(cur_text, is_ok)


# ━━━━━━━ 主窗口 ━━━━━━━
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.b_file = ""
        self.rule_file = ""
        self.mapping_file = ""
        self.mapping_data = None
        self.result = None
        self.all_table_rows = []
        self._init_ui()
        self._load_saved()

    def _load_saved(self):
        global _font_scale
        cfg = load_config()
        scale = cfg.get("font_scale", 100)
        _font_scale = max(0.8, min(2.0, scale / 100.0))
        self._apply_all_styles()

        # 从缓存目录加载规则表
        cached_rule = os.path.join(CACHE_DIR, "rule_table.xlsx")
        if os.path.exists(cached_rule):
            self.rule_file = cached_rule
            self.rule_block.drop.set_loaded("rule_table.xlsx")
            self.rule_block.set_status("✅ 已加载（缓存）", True)

        # 从缓存目录加载费用配置表
        cached_mapping = os.path.join(CACHE_DIR, "mapping_table.xlsx")
        if os.path.exists(cached_mapping):
            self.mapping_file = cached_mapping
            self.mapping_block.drop.set_loaded("mapping_table.xlsx")
        cached = load_mapping_cache()
        if cached:
            self.mapping_data = cached
            cnt = sum(len(v) for v in cached.values())
            self.mapping_block.set_status(
                f"✅ 已加载（缓存） {len(cached)} 个科目 {cnt} 条规则", True
            )

        # B表不缓存，仍从config加载
        bp = cfg.get("b_file", "")
        if bp and os.path.exists(bp):
            self.b_file = bp
            self.b_block.drop.set_loaded(os.path.basename(bp))

    def _save_paths(self):
        cfg = load_config()
        if self.rule_file:
            cfg["rule_file"] = self.rule_file
        if self.b_file:
            cfg["b_file"] = self.b_file
        if self.mapping_file:
            cfg["mapping_file"] = self.mapping_file
        save_config(cfg)

    def _init_ui(self):
        self.setWindowTitle(f"{APP_NAME}  {VERSION}")
        self.setMinimumSize(1200, 850)
        self.resize(1400, 960)
        self.showMaximized()

        self.setStyleSheet(f"""
            QMainWindow {{
                background: {C_BG};
            }}
            QStatusBar {{
                background: {C_CARD};
                color: {C_TEXT_SEC};
                font-size: {FONT_SMALL()}px;
                border-top: 1px solid {C_BORDER};
            }}
        """)
        QApplication.setFont(QFont(FONT_FAMILY, FONT_BODY()))

        # ── 菜单栏 ──
        menubar = self.menuBar()
        menubar.setStyleSheet(f"""
            QMenuBar {{
                background: {C_CARD};
                border-bottom: 1px solid {C_BORDER};
                padding: 2px 8px;
                font-size: {FONT_SMALL()}px;
                color: {C_TEXT_SEC};
            }}
            QMenuBar::item {{
                padding: 4px 10px;
                border-radius: 4px;
            }}
            QMenuBar::item:selected {{
                background: {C_BORDER_L};
                color: {C_TEXT};
            }}
        """)
        font_menu = menubar.addMenu("字体(&T)")
        font_action = QAction("字体大小设置...", self)
        font_action.triggered.connect(self._show_font_dialog)
        font_menu.addAction(font_action)

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setSpacing(0)
        root.setContentsMargins(0, 0, 0, 0)

        # ── 标题栏 ──
        title_bar = QFrame()
        title_bar.setStyleSheet(f"""
            QFrame {{
                background: {C_CARD};
                border-bottom: 1px solid {C_BORDER};
            }}
        """)
        tb_ly = QVBoxLayout(title_bar)
        tb_ly.setContentsMargins(32, 16, 32, 0)
        tb_ly.setSpacing(0)

        # 标题行
        title_row = QHBoxLayout()
        self.hdr_title = QLabel(f"{APP_NAME}")
        self.hdr_title.setStyleSheet(f"""
            font-size: {FONT_TITLE()}px; font-weight: bold;
            color: {C_TEXT}; background: transparent;
        """)
        self.hdr_version = QLabel(VERSION)
        self.hdr_version.setStyleSheet(f"""
            font-size: {FONT_SMALL()}px; font-weight: 500;
            color: {C_TEXT_AUX}; background: transparent;
            padding: 2px 8px;
        """)
        title_row.addWidget(self.hdr_title)
        title_row.addSpacing(12)
        title_row.addWidget(self.hdr_version)
        title_row.addStretch()
        tb_ly.addLayout(title_row)
        tb_ly.addSpacing(16)

        # ── Tab 栏 ──
        self.page_tab = QTabBar()
        self.page_tab.setDrawBase(False)
        self.page_tab.setStyleSheet(f"""
            QTabBar::tab {{
                background: transparent;
                border: none;
                border-bottom: 2px solid transparent;
                padding: 10px 24px;
                font-size: {FONT_H2()}px;
                font-weight: 500;
                color: {C_TEXT_SEC};
            }}
            QTabBar::tab:selected {{
                color: {C_PRIMARY};
                border-bottom: 2px solid {C_PRIMARY};
                font-weight: 600;
            }}
            QTabBar::tab:hover {{
                color: {C_TEXT};
            }}
        """)
        self.page_tab.addTab("📄 文件校验")
        self.page_tab.addTab("📋 表格校验")
        self.page_tab.currentChanged.connect(lambda idx: self.stack.setCurrentIndex(idx))
        tb_ly.addWidget(self.page_tab)

        self.title_bar_frame = title_bar
        root.addWidget(title_bar)

        # ── 内容区 ──
        content = QWidget()
        content.setStyleSheet(f"background: {C_BG};")
        content_ly = QVBoxLayout(content)
        content_ly.setContentsMargins(32, 24, 32, 24)
        content_ly.setSpacing(0)

        self.stack = QStackedWidget()
        self.stack.addWidget(self._build_file_page())
        self.stack.addWidget(self._build_grid_page())
        content_ly.addWidget(self.stack, 1)
        root.addWidget(content, 1)

        self.statusBar().showMessage(f"  {APP_NAME} {VERSION}  |  就绪")

    # ════════════════ 页面一：文件校验 ════════════════
    def _build_file_page(self):
        page = QWidget()
        page.setStyleSheet(f"background: {C_BG};")
        root = QVBoxLayout(page)
        root.setSpacing(20)
        root.setContentsMargins(0, 0, 0, 0)

        # ── 上方：左右分区 ──
        upload_row = QHBoxLayout()
        upload_row.setSpacing(20)

        # 左侧：B表上传
        self.b_block = FileBlock(
            "📄 B表（Excel文件上传）",
            drop_text="📁 将B表Excel文件拖拽到此处\n或点击上方按钮选择",
            drop_h=160,
            show_reload=True
        )
        self.b_block.file_chosen.connect(self._on_b_chosen)
        if self.b_block.reload_btn:
            self.b_block.reload_btn.clicked.connect(self._reload_b)
        upload_row.addWidget(self.b_block, 1)

        # 右侧：两个校验表竖排
        right_col = QVBoxLayout()
        right_col.setSpacing(12)

        self.rule_block = FileBlock(
            "📋 校验规则表",
            subtitle="记住路径",
            drop_text="将规则表拖到此处",
            drop_h=60,
            show_reload=True
        )
        self.rule_block.file_chosen.connect(self._on_rule_chosen)
        if self.rule_block.reload_btn:
            self.rule_block.reload_btn.clicked.connect(self._reload_rule)
        right_col.addWidget(self.rule_block)

        self.mapping_block = FileBlock(
            "📊 费用配置表",
            subtitle="记住并缓存",
            drop_text="将费用配置表拖到此处",
            drop_h=60,
            show_reload=False
        )
        self.mapping_block.file_chosen.connect(self._on_mapping_chosen)
        right_col.addWidget(self.mapping_block)

        right_widget = QWidget()
        right_widget.setLayout(right_col)
        upload_row.addWidget(right_widget, 1)

        root.addLayout(upload_row)

        # ── 操作按钮行 ──
        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)
        self.vbtn = QPushButton("🔍 开始校验")
        self.vbtn.setFixedHeight(42)
        self.vbtn.setCursor(Qt.PointingHandCursor)
        self.vbtn.setStyleSheet(_orange_btn_style())
        self.vbtn.clicked.connect(self._start)
        self.dbtn = QPushButton("⬇️ 下载报告")
        self.dbtn.setFixedHeight(42)
        self.dbtn.setCursor(Qt.PointingHandCursor)
        self.dbtn.setStyleSheet(_secondary_btn_style())
        self.dbtn.setEnabled(False)
        self.dbtn.clicked.connect(self._download)
        btn_row.addWidget(self.vbtn)
        btn_row.addWidget(self.dbtn)
        btn_row.addStretch()
        root.addLayout(btn_row)

        # ── 进度条 ──
        self.prog = QProgressBar()
        self.prog.setFixedHeight(6)
        self.prog.setValue(0)
        self.prog.setTextVisible(False)
        self.prog.setStyleSheet(f"""
            QProgressBar {{
                background: {C_BORDER};
                border-radius: 3px;
                border: none;
            }}
            QProgressBar::chunk {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {C_PRIMARY}, stop:1 {C_PRIMARY});
                border-radius: 3px;
            }}
        """)
        prog_wrap = QVBoxLayout()
        prog_wrap.setSpacing(4)
        prog_wrap.addWidget(self.prog)
        self.prog_text = QLabel("")
        self.prog_text.setAlignment(Qt.AlignRight)
        self.prog_text.setStyleSheet(f"""
            font-size: {FONT_SMALL()}px; color: {C_TEXT_AUX};
            background: transparent;
        """)
        prog_wrap.addWidget(self.prog_text)
        root.addLayout(prog_wrap)

        # ── 统计卡片行 ──
        self.stat_row = StatRow()
        root.addWidget(self.stat_row)

        # ── 校验结果区 ──
        result_card = QFrame()
        result_card.setStyleSheet(f"""
            QFrame#resultCard {{
                background: {C_CARD};
                border: 1px solid {C_BORDER};
                border-radius: 10px;
            }}
        """)
        result_card.setObjectName("resultCard")
        result_ly = QVBoxLayout(result_card)
        result_ly.setContentsMargins(24, 20, 24, 20)
        result_ly.setSpacing(12)

        result_title = QLabel("📋 校验结果")
        result_title.setStyleSheet(f"""
            font-size: {FONT_H2()}px; font-weight: bold;
            color: {C_TEXT}; background: transparent; border: none;
        """)
        result_ly.addWidget(result_title)
        self._result_title_lbl = result_title

        # 筛选 Tab
        self.filter_tabs = QTabBar()
        self.filter_tabs.setDrawBase(False)
        self.filter_tabs.setStyleSheet(self._filter_tab_style())
        for t in ["全部", "必输项", "借贷平衡", "记账码", "禁填字段", "费用类别", "⚠️ 警告"]:
            self.filter_tabs.addTab(t)
        self.filter_tabs.currentChanged.connect(self._tab_changed)
        result_ly.addWidget(self.filter_tabs)

        # 结果表格
        self.tbl = QTableWidget()
        self.tbl.setColumnCount(4)
        self.tbl.setHorizontalHeaderLabels(["行号", "总账科目", "错误类型", "错误详情"])
        self.tbl.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.tbl.setColumnWidth(0, 60)
        self.tbl.setColumnWidth(1, 200)
        self.tbl.setColumnWidth(2, 260)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl.verticalHeader().setDefaultSectionSize(36)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setStyleSheet(self._result_table_style())
        result_ly.addWidget(self.tbl, 1)

        root.addWidget(result_card, 1)
        return page

    def _filter_tab_style(self):
        return f"""
            QTabBar::tab {{
                background: {C_BORDER_L};
                border: 1px solid transparent;
                border-radius: 6px;
                padding: 6px 14px;
                margin-right: 4px;
                font-size: {FONT_GRID_H()}px;
                font-weight: 500;
                color: {C_TEXT_SEC};
            }}
            QTabBar::tab:selected {{
                background: {C_PRIMARY_L};
                color: {C_PRIMARY};
                border-color: #FFB74D;
                font-weight: 600;
            }}
            QTabBar::tab:hover {{
                background: {C_BORDER};
                color: {C_TEXT};
            }}
        """

    def _result_table_style(self):
        return f"""
            QTableWidget {{
                border: 1px solid {C_BORDER};
                border-radius: 8px;
                gridline-color: {C_BORDER_L};
                font-size: {FONT_BODY()}px;
                background: {C_CARD};
                alternate-background-color: {C_BORDER_L};
            }}
            QHeaderView::section {{
                background: {C_BG};
                color: {C_TEXT_SEC};
                font-weight: 600;
                padding: 10px 16px;
                border: none;
                border-bottom: 1px solid {C_BORDER};
                font-size: {FONT_GRID_H()}px;
            }}
            QTableWidget::item {{
                padding: 10px 16px;
            }}
            QTableWidget::item:selected {{
                background: #FF8C00;
                color: #1C1917;
            }}
        """

    # ════════════════ 页面二：表格校验 ════════════════
    def _build_grid_page(self):
        page = QWidget()
        page.setStyleSheet(f"background: {C_BG};")
        root = QVBoxLayout(page)
        root.setSpacing(16)
        root.setContentsMargins(0, 0, 0, 0)

        # ── 提示信息栏 ──
        self.tip_label = QLabel(
            "💡 从 Excel 或 SAP 复制数据后，按 Ctrl + V 粘贴到下方表格"
        )
        self.tip_label.setStyleSheet(f"""
            QLabel {{
                color: {C_ACCENT};
                font-size: {FONT_BODY()}px;
                background: {C_DROP_BG};
                padding: 14px 20px;
                border-radius: 8px;
                border: 1px solid {C_DROP_BD};
            }}
        """)
        root.addWidget(self.tip_label)

        # ── 操作按钮行 ──
        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)
        self.g_vbtn = QPushButton("🔍 校验")
        self.g_vbtn.setFixedHeight(42)
        self.g_vbtn.setCursor(Qt.PointingHandCursor)
        self.g_vbtn.setStyleSheet(_orange_btn_style())
        self.g_vbtn.clicked.connect(self._grid_validate)

        self.g_clr = QPushButton("🗑️ 清空数据")
        self.g_clr.setFixedHeight(42)
        self.g_clr.setCursor(Qt.PointingHandCursor)
        self.g_clr.setStyleSheet(_danger_btn_style())
        self.g_clr.clicked.connect(self._grid_clear)

        self.g_dl = QPushButton("⬇️ 下载报告")
        self.g_dl.setFixedHeight(42)
        self.g_dl.setCursor(Qt.PointingHandCursor)
        self.g_dl.setStyleSheet(_secondary_btn_style())
        self.g_dl.setEnabled(False)
        self.g_dl.clicked.connect(self._grid_download)

        btn_row.addWidget(self.g_vbtn)
        btn_row.addWidget(self.g_clr)
        btn_row.addWidget(self.g_dl)
        btn_row.addStretch()
        root.addLayout(btn_row)

        # ── 进度条 ──
        self.g_prog = QProgressBar()
        self.g_prog.setFixedHeight(6)
        self.g_prog.setValue(0)
        self.g_prog.setTextVisible(False)
        self.g_prog.setStyleSheet(f"""
            QProgressBar {{
                background: {C_BORDER};
                border-radius: 3px;
                border: none;
            }}
            QProgressBar::chunk {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {C_PRIMARY}, stop:1 {C_PRIMARY});
                border-radius: 3px;
            }}
        """)
        root.addWidget(self.g_prog)

        # ── 统计卡片行 ──
        self.g_stat_row = StatRow()
        root.addWidget(self.g_stat_row)

        # ── 大表格 ──
        nc = NUM_COLS + 1
        self.grid = PasteableTable()
        self.grid.setColumnCount(nc)
        self.grid.setHorizontalHeaderLabels(["校验结果"] + COL_LETTERS)
        self.grid.setRowCount(GRID_HEADER_ROWS + 200)

        # 填充表头（只保留中文列名1行）
        # 校验结果列（第0列）
        item = QTableWidgetItem("校验结果")
        item.setBackground(QColor("#FFFFFF"))
        item.setForeground(QColor("#8B2500"))
        item.setFont(QFont(FONT_FAMILY, FONT_SMALL(), QFont.Bold))
        item.setFlags(Qt.ItemIsEnabled)
        item.setTextAlignment(Qt.AlignCenter)
        self.grid.setItem(0, RESULT_COL, item)
        # 数据列（从第1列开始）
        for c, v in enumerate(HEADER_ROW1):
            item = QTableWidgetItem(str(v) if v is not None else "")
            item.setBackground(QColor("#FFFFFF"))
            item.setForeground(QColor("#8B2500"))
            item.setFont(QFont(FONT_FAMILY, FONT_SMALL(), QFont.Bold))
            item.setFlags(Qt.ItemIsEnabled)
            item.setTextAlignment(Qt.AlignCenter)
            self.grid.setItem(0, c + DATA_COL_OFFSET, item)

        self.grid.setStyleSheet(self._grid_table_style())
        self.grid.horizontalHeader().setDefaultSectionSize(100)
        self.grid.setColumnWidth(RESULT_COL, 360)
        self.grid.verticalHeader().setDefaultSectionSize(30)
        self.grid.verticalHeader().setVisible(False)
        self.grid.setMouseTracking(True)
        self.grid.cellEntered.connect(self._grid_cell_hover)
        root.addWidget(self.grid, 1)

        self.grid_errors = {}
        self.grid_result = None
        return page

    def _grid_table_style(self):
        return f"""
            QTableWidget {{
                border: 1px solid {C_BORDER};
                border-radius: 8px;
                gridline-color: #F0EEEB;
                font-size: {FONT_GRID()}px;
                background: {C_CARD};
            }}
            QHeaderView::section {{
                background: {C_CARD};
                color: #8B2500;
                font-weight: 600;
                padding: 6px;
                border: none;
                border-bottom: 2px solid #D4A574;
                font-size: {FONT_BODY()}px;
            }}
            QTableWidget::item:selected {{
                background: #FFD8A6;
                color: #1C1917;
            }}
        """

    # ════════════════ 字体缩放 ════════════════
    def _show_font_dialog(self):
        global _font_scale
        dlg = QDialog(self)
        dlg.setWindowTitle("字体大小设置")
        dlg.setFixedSize(420, 280)
        dlg.setStyleSheet(f"background: {C_CARD};")
        layout = QVBoxLayout(dlg)
        layout.setSpacing(12)

        current_val = int(_font_scale * 100)
        val_label = QLabel(f"当前：{current_val}%")
        val_label.setAlignment(Qt.AlignCenter)
        val_label.setStyleSheet(f"""
            font-size: {FONT_H2()}px; font-weight: bold;
            color: {C_TEXT};
        """)
        layout.addWidget(val_label)

        slider_row = QHBoxLayout()
        min_label = QLabel("80%")
        min_label.setStyleSheet(f"font-size: {FONT_SMALL()}px; color: {C_TEXT_SEC};")
        max_label = QLabel("200%")
        max_label.setStyleSheet(f"font-size: {FONT_SMALL()}px; color: {C_TEXT_SEC};")
        slider = QSlider(Qt.Horizontal)
        slider.setRange(80, 200)
        slider.setSingleStep(10)
        slider.setPageStep(10)
        slider.setValue(current_val)
        slider.setStyleSheet(f"""
            QSlider::groove:horizontal {{
                border: none;
                height: 6px;
                background: {C_BORDER};
                border-radius: 3px;
            }}
            QSlider::handle:horizontal {{
                background: {C_PRIMARY};
                border: none;
                width: 18px;
                margin: -6px 0;
                border-radius: 9px;
            }}
            QSlider::sub-page:horizontal {{
                background: {C_PRIMARY};
                border-radius: 3px;
            }}
        """)
        slider_row.addWidget(min_label)
        slider_row.addWidget(slider, 1)
        slider_row.addWidget(max_label)
        layout.addLayout(slider_row)

        preview = QLabel("这是预览文字 ABC 123")
        preview.setAlignment(Qt.AlignCenter)
        preview.setStyleSheet(f"""
            QLabel {{
                font-size: {fs(BASE_BODY)}px; color: {C_TEXT};
                background: {C_BG}; border: 1px solid {C_BORDER};
                border-radius: 6px; padding: 12px;
            }}
        """)
        layout.addWidget(preview)

        def on_slider_change(val):
            snapped = round(val / 10) * 10
            if slider.value() != snapped:
                slider.setValue(snapped)
                return
            val_label.setText(f"当前：{snapped}%")
            preview_size = max(8, int(BASE_BODY * snapped / 100.0 + 0.5))
            preview.setStyleSheet(f"""
                QLabel {{
                    font-size: {preview_size}px; color: {C_TEXT};
                    background: {C_BG}; border: 1px solid {C_BORDER};
                    border-radius: 6px; padding: 12px;
                }}
            """)
        slider.valueChanged.connect(on_slider_change)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.button(QDialogButtonBox.Ok).setText("确定")
        btn_box.button(QDialogButtonBox.Cancel).setText("取消")
        btn_box.button(QDialogButtonBox.Ok).setStyleSheet(_orange_btn_style())
        btn_box.accepted.connect(dlg.accept)
        btn_box.rejected.connect(dlg.reject)
        layout.addWidget(btn_box)

        if dlg.exec_() == QDialog.Accepted:
            new_scale = round(slider.value() / 10) * 10
            _font_scale = new_scale / 100.0
            self._apply_all_styles()
            cfg = load_config()
            cfg["font_scale"] = int(_font_scale * 100)
            save_config(cfg)

    def _apply_all_styles(self):
        """重新应用所有样式表（字体缩放后调用）"""
        global _font_scale
        QApplication.setFont(QFont(FONT_FAMILY, FONT_BODY()))

        # 主窗口
        self.setStyleSheet(f"""
            QMainWindow {{
                background: {C_BG};
            }}
            QStatusBar {{
                background: {C_CARD};
                color: {C_TEXT_SEC};
                font-size: {FONT_SMALL()}px;
                border-top: 1px solid {C_BORDER};
            }}
        """)

        # 菜单栏
        self.menuBar().setStyleSheet(f"""
            QMenuBar {{
                background: {C_CARD};
                border-bottom: 1px solid {C_BORDER};
                padding: 2px 8px;
                font-size: {FONT_SMALL()}px;
                color: {C_TEXT_SEC};
            }}
            QMenuBar::item {{
                padding: 4px 10px;
                border-radius: 4px;
            }}
            QMenuBar::item:selected {{
                background: {C_BORDER_L};
                color: {C_TEXT};
            }}
        """)

        # 标题
        self.hdr_title.setStyleSheet(f"""
            font-size: {FONT_TITLE()}px; font-weight: bold;
            color: {C_TEXT}; background: transparent;
        """)
        self.hdr_version.setStyleSheet(f"""
            font-size: {FONT_SMALL()}px; font-weight: 500;
            color: {C_TEXT_AUX}; background: transparent;
            padding: 2px 8px;
        """)

        # 页面 Tab
        self.page_tab.setStyleSheet(f"""
            QTabBar::tab {{
                background: transparent;
                border: none;
                border-bottom: 2px solid transparent;
                padding: 10px 24px;
                font-size: {FONT_H2()}px;
                font-weight: 500;
                color: {C_TEXT_SEC};
            }}
            QTabBar::tab:selected {{
                color: {C_PRIMARY};
                border-bottom: 2px solid {C_PRIMARY};
                font-weight: 600;
            }}
            QTabBar::tab:hover {{
                color: {C_TEXT};
            }}
        """)

        # 文件上传卡片
        self.b_block.refresh_style()
        self.rule_block.refresh_style()
        self.mapping_block.refresh_style()

        # 校验/下载按钮（文件校验页）
        btn_h = max(36, int(42 * _font_scale + 0.5))
        self.vbtn.setFixedHeight(btn_h)
        self.vbtn.setStyleSheet(_orange_btn_style())
        self.dbtn.setFixedHeight(btn_h)
        self.dbtn.setStyleSheet(_secondary_btn_style())

        # 进度条文字
        self.prog_text.setStyleSheet(f"""
            font-size: {FONT_SMALL()}px; color: {C_TEXT_AUX};
            background: transparent;
        """)

        # 统计卡片
        self.stat_row.refresh_style()
        self.g_stat_row.refresh_style()

        # 筛选 Tab
        self.filter_tabs.setStyleSheet(self._filter_tab_style())

        # 结果标题
        self._result_title_lbl.setStyleSheet(f"""
            font-size: {FONT_H2()}px; font-weight: bold;
            color: {C_TEXT}; background: transparent; border: none;
        """)

        # 结果表格
        row_h = max(28, int(36 * _font_scale + 0.5))
        self.tbl.verticalHeader().setDefaultSectionSize(row_h)
        self.tbl.setStyleSheet(self._result_table_style())

        # 提示信息（Tab2）
        self.tip_label.setStyleSheet(f"""
            QLabel {{
                color: {C_ACCENT};
                font-size: {FONT_BODY()}px;
                background: {C_DROP_BG};
                padding: 14px 20px;
                border-radius: 8px;
                border: 1px solid {C_DROP_BD};
            }}
        """)

        # 表格校验页按钮
        self.g_vbtn.setFixedHeight(btn_h)
        self.g_vbtn.setStyleSheet(_orange_btn_style())
        self.g_clr.setFixedHeight(btn_h)
        self.g_clr.setStyleSheet(_danger_btn_style())
        self.g_dl.setFixedHeight(btn_h)
        self.g_dl.setStyleSheet(_secondary_btn_style())

        # grid 表格
        grid_row_h = max(24, int(30 * _font_scale + 0.5))
        grid_col_w = max(80, int(100 * _font_scale + 0.5))
        grid_err_w = max(240, int(360 * _font_scale + 0.5))
        self.grid.verticalHeader().setDefaultSectionSize(grid_row_h)
        self.grid.horizontalHeader().setDefaultSectionSize(grid_col_w)
        self.grid.setColumnWidth(RESULT_COL, grid_err_w)
        self.grid.setStyleSheet(self._grid_table_style())

        # 更新 grid 表头行字体
        for r in range(GRID_HEADER_ROWS):
            for c in range(NUM_COLS + 1):
                item = self.grid.item(r, c)
                if item:
                    item.setFont(QFont(FONT_FAMILY, FONT_SMALL(), QFont.Bold))

    # ════════════════ Tab2 事件 ════════════════
    def _grid_cell_hover(self, row, col):
        data_col = col - DATA_COL_OFFSET
        if 0 <= data_col < NUM_COLS:
            cl = COL_LETTERS[data_col]
            key = (row, cl)
            if key in self.grid_errors:
                pos = self.grid.viewport().mapToGlobal(QPoint(0, 0))
                QToolTip.showText(pos, self.grid_errors[key])

    def _grid_clear(self):
        for r in range(GRID_HEADER_ROWS, self.grid.rowCount()):
            for c in range(self.grid.columnCount()):
                item = self.grid.item(r, c)
                if item:
                    item.setText("")
                    item.setBackground(QColor("white"))
                    item.setForeground(QColor(C_TEXT))
                    item.setToolTip("")
                else:
                    self.grid.setItem(r, c, QTableWidgetItem(""))
        self.grid_errors = {}
        self.grid_result = None
        self.g_dl.setEnabled(False)
        self.g_stat_row.set_values()

    def _grid_validate(self):
        if not self.rule_file:
            QMessageBox.warning(self, "提示", "请先在「文件校验」页面上传校验规则表！")
            return
        all_rows = {}
        for r in range(GRID_HEADER_ROWS, self.grid.rowCount()):
            row_data = []
            has_data = False
            for c in range(NUM_COLS):
                item = self.grid.item(r, c + DATA_COL_OFFSET)
                v = item.text().strip() if item else ""
                if v:
                    has_data = True
                # 金额列(O列)和记账码列(J列)保持字符串，避免精度问题和前导零丢失
                if c == O_IDX:
                    # 去除千分位逗号
                    v_clean = v.replace(',', '') if v else v
                    row_data.append(v_clean if v_clean != "" else None)
                elif c == J_IDX:
                    # 记账码保持字符串，保留前导零
                    row_data.append(v if v != "" else None)
                else:
                    try:
                        v = int(v)
                    except:
                        try:
                            v = float(v)
                        except:
                            pass
                    row_data.append(v if v != "" else None)
            if has_data:
                all_rows[r + 1] = row_data
        

        if not all_rows:
            QMessageBox.warning(self, "提示", "表格中没有数据，请先粘贴数据！")
            return
        self.g_vbtn.setEnabled(False)
        self.g_prog.setValue(0)
        self.g_worker = GridValidateWorker(all_rows, self.rule_file, self.mapping_data)
        self.g_worker.progress.connect(self.g_prog.setValue)
        self.g_worker.finished.connect(self._grid_done)
        self.g_worker.error.connect(self._grid_err)
        self.g_worker.start()

    def _grid_done(self, result):
        errors = result["errors"]
        warnings = result["warnings"]
        total = result["total"]
        ec = len(errors)
        wc = len([w for w in warnings if w[2] == '科目不在规则表'])
        ok = max(0, total - ec - wc)
        self.g_stat_row.set_values(total, ok, ec, wc)

        # 清除之前标色
        self.grid_errors = {}
        for r in range(GRID_HEADER_ROWS, self.grid.rowCount()):
            for c in range(self.grid.columnCount()):
                item = self.grid.item(r, c)
                if item:
                    item.setBackground(QColor("white"))
                    item.setForeground(QColor(C_TEXT))
                    item.setToolTip("")

        ecells = set()
        for r, el in errors.items():
            for _, cols, _ in el:
                for cl in cols:
                    ecells.add((r, cl))
        wa = {w[0]: w[1] for w in warnings if w[2] == '科目不在规则表'}

        for r, el in errors.items():
            gr = r - 1
            if gr >= self.grid.rowCount():
                continue
            for c in range(NUM_COLS):
                item = self.grid.item(gr, c + DATA_COL_OFFSET)
                if not item:
                    continue
                cl = COL_LETTERS[c]
                if (r, cl) in ecells:
                    item.setBackground(QColor("#FDE68A"))
                    item.setForeground(QColor(C_ACCENT))
                    msgs = [e[2] for e in el if cl in e[1]]
                    if msgs:
                        tip = "\n".join(msgs)
                        item.setToolTip(tip)
                        self.grid_errors[(gr, cl)] = tip
                else:
                    item.setBackground(QColor("#FEF3C7"))
            ei = self.grid.item(gr, RESULT_COL)
            if not ei:
                ei = QTableWidgetItem()
                self.grid.setItem(gr, RESULT_COL, ei)
            ei.setText(" | ".join([e[2] for e in el]))
            ei.setForeground(QColor(C_ERROR))
            ei.setBackground(QColor("#FEF3C7"))

        for r, acct in wa.items():
            gr = r - 1
            if gr >= self.grid.rowCount():
                continue
            for c in range(NUM_COLS):
                item = self.grid.item(gr, c + DATA_COL_OFFSET)
                if item:
                    item.setBackground(QColor("#FFE0B2"))
            ei = self.grid.item(gr, RESULT_COL)
            if not ei:
                ei = QTableWidgetItem()
                self.grid.setItem(gr, RESULT_COL, ei)
            ei.setText(f"⚠️ 科目{acct}不在规则表中")
            ei.setForeground(QColor(C_WARN))
            ei.setBackground(QColor("#FFE0B2"))

        self.grid_result = result
        self.g_vbtn.setEnabled(True)
        self.g_dl.setEnabled(True)
        self.g_prog.setValue(100)
        self.statusBar().showMessage(
            f"  表格校验完成：{ec}个错误 {wc}个警告  {datetime.now().strftime('%H:%M:%S')}"
        )

    def _grid_err(self, msg):
        self.g_vbtn.setEnabled(True)
        self.g_prog.setValue(0)
        QMessageBox.critical(self, "校验出错", f"错误：\n{msg}")

    def _grid_download(self):
        if not self.grid_result:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        for r, hrow in enumerate([HEADER_ROW1, HEADER_ROW2, HEADER_ROW3], 1):
            for c, v in enumerate(hrow, 1):
                ws.cell(row=r, column=c, value=v)
        for r in range(GRID_HEADER_ROWS, self.grid.rowCount()):
            has = False
            for c in range(NUM_COLS):
                item = self.grid.item(r, c + DATA_COL_OFFSET)
                v = item.text().strip() if item else ""
                if v:
                    has = True
                try:
                    v = int(v)
                except:
                    try:
                        v = float(v)
                    except:
                        pass
                ws.cell(row=r + 3, column=c + 1, value=v if v != "" else None)
            if not has:
                break
        rm, _ = load_rule_table(self.rule_file)
        out_wb = build_report(
            self.grid_result["errors"],
            self.grid_result["warnings"],
            wb, ws, rm,
            self.grid_result["total"]
        )
        dn = f"校验报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        sp, _ = QFileDialog.getSaveFileName(self, "保存", dn, "Excel (*.xlsx)")
        if sp:
            try:
                out_wb.save(sp)
                QMessageBox.information(self, "✅", f"报告已保存到：\n{sp}")
            except Exception as e:
                QMessageBox.critical(self, "失败", str(e))

    # ════════════════ 文件校验公共方法 ════════════════
    def _on_b_chosen(self, p):
        self.b_file = p
        self.b_block.drop.set_loaded(os.path.basename(p))
        self._save_paths()

    def _on_rule_chosen(self, p):
        self.rule_file = p
        shutil.copy2(p, os.path.join(CACHE_DIR, "rule_table.xlsx"))
        self.rule_block.drop.set_loaded(os.path.basename(p))
        self.rule_block.set_status("✔ 已加载 · 已缓存", True)
        self._save_paths()

    def _on_mapping_chosen(self, p):
        self._load_mapping(p)

    def _reload_b(self):
        if self.b_file and os.path.exists(self.b_file):
            self.b_block.drop.set_loaded(os.path.basename(self.b_file))
            self.b_block.set_status("✔ 已重新加载", True)
        else:
            QMessageBox.warning(self, "提示", "请先选择B表文件！")

    def _reload_rule(self):
        if self.rule_file and os.path.exists(self.rule_file):
            self.rule_block.drop.set_loaded(os.path.basename(self.rule_file))
            self.rule_block.set_status("✔ 已重新加载 · 路径已记住", True)
        else:
            QMessageBox.warning(self, "提示", "请先选择校验规则表文件！")

    def _load_mapping(self, path):
        try:
            data = load_mapping_table(path)
            self.mapping_data = data
            self.mapping_file = path
            shutil.copy2(path, os.path.join(CACHE_DIR, "mapping_table.xlsx"))
            save_mapping_cache(data)
            self._save_paths()
            self.mapping_block.drop.set_loaded(os.path.basename(path))
            cnt = sum(len(v) for v in data.values())
            self.mapping_block.set_status(
                f"✔ 已加载并缓存：{len(data)} 个科目 {cnt} 条规则", True
            )
        except Exception as ex:
            QMessageBox.critical(self, "加载失败", str(ex))

    def _start(self):
        if not self.b_file:
            QMessageBox.warning(self, "提示", "请先上传B表！")
            return
        if not self.rule_file:
            QMessageBox.warning(self, "提示", "请先上传校验规则表！")
            return
        self.vbtn.setEnabled(False)
        self.dbtn.setEnabled(False)
        self.prog.setValue(0)
        self.prog_text.setText("")
        self.tbl.setRowCount(0)
        self.stat_row.set_values("...", "...", "...", "...")
        self.worker = ValidateWorker(self.b_file, self.rule_file, self.mapping_data)
        self.worker.progress.connect(self.prog.setValue)
        self.worker.finished.connect(self._done)
        self.worker.error.connect(self._err)
        self.worker.start()

    def _done(self, result):
        self.result = result
        errs = result["errors"]
        warns = result["warnings"]
        total = result["total"]
        ec = len(errs)
        wc = len([w for w in warns if w[2] == '科目不在规则表'])
        ok = max(0, total - ec - wc)
        self.stat_row.set_values(total, ok, ec, wc)
        self.prog_text.setText(f"校验完成 — 共 {total} 行已处理")

        self.all_table_rows = []
        ws = result.get("ws")
        for r in sorted(errs.keys()):
            acct = ''
            if ws:
                cell_val = ws.cell(row=r, column=13).value
                if cell_val is not None:
                    acct = str(int(cell_val)) if isinstance(cell_val, float) else str(cell_val).strip()
            for et, cols, msg in errs[r]:
                if '必输' in et or '金额' in et:
                    cat = '必输项'
                elif '借贷' in et:
                    cat = '借贷平衡'
                elif '记账码' in et:
                    cat = '记账码'
                elif '禁填' in et:
                    cat = '禁填字段'
                elif '费用' in et:
                    cat = '费用类别'
                else:
                    cat = '其他'
                self.all_table_rows.append((str(r), acct, f"❌ {et}", msg, cat, False))
        for w in warns:
            if w[2] == '科目不在规则表':
                self.all_table_rows.append(
                    (str(w[0]), w[1], "⚠️ 科目不在规则表", "请核实", '警告', True)
                )
            elif '记账码' in w[2]:
                self.all_table_rows.append(
                    (str(w[0]), '', f"⚠️ {w[2]}", str(w[1]), '记账码', True)
                )
        self.filter_tabs.setCurrentIndex(0)
        self._fill(self.all_table_rows)
        self.vbtn.setEnabled(True)
        self.dbtn.setEnabled(True)
        self.statusBar().showMessage(
            f"  校验完成  {datetime.now().strftime('%H:%M:%S')}"
        )

    def _tab_changed(self, idx):
        if not self.all_table_rows:
            return
        fmap = {
            0: None, 1: '必输项', 2: '借贷平衡', 3: '记账码',
            4: '禁填字段', 5: '费用类别', 6: '警告'
        }
        f = fmap.get(idx)
        if f is None:
            self._fill(self.all_table_rows)
        else:
            self._fill([r for r in self.all_table_rows if r[4] == f or (f == '警告' and r[5])])

    def _fill(self, rows):
        self.tbl.setRowCount(len(rows))
        for i, (rno, acct, et, detail, cat, iw) in enumerate(rows):
            for j, v in enumerate([rno, acct, et, detail]):
                item = QTableWidgetItem(v)
                if j == 0:
                    item.setForeground(QColor(C_TEXT_SEC))
                    item.setFont(QFont(FONT_FAMILY, FONT_GRID()))
                elif j == 1:
                    item.setForeground(QColor(C_TEXT))
                    item.setFont(QFont("Consolas", FONT_GRID()))
                elif j == 2:
                    if iw:
                        item.setForeground(QColor(C_WARN))
                    else:
                        item.setForeground(QColor(C_ERROR))
                    item.setFont(QFont(FONT_FAMILY, FONT_GRID()))
                elif j == 3:
                    if iw:
                        item.setForeground(QColor(C_ACCENT))
                    else:
                        item.setForeground(QColor(C_TEXT_SEC))
                    item.setFont(QFont(FONT_FAMILY, FONT_GRID()))
                self.tbl.setItem(i, j, item)

    def _err(self, msg):
        self.vbtn.setEnabled(True)
        self.prog.setValue(0)
        QMessageBox.critical(self, "校验出错", f"错误：\n{msg}")

    def _download(self):
        if not self.result:
            return
        dn = f"校验报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        sp, _ = QFileDialog.getSaveFileName(self, "保存", dn, "Excel (*.xlsx)")
        if sp:
            try:
                self.result["out_wb"].save(sp)
                QMessageBox.information(self, "✅", f"报告已保存到：\n{sp}")
            except Exception as e:
                QMessageBox.critical(self, "失败", str(e))


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
